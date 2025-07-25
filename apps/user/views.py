import base64
import json
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.shortcuts import HttpResponse
from rest_framework.viewsets import ViewSet
from rest_framework_jwt.settings import api_settings
from django.db import transaction
from service_error.common import COMMON_RERROR
from service_error.user import USER_RERROR
from apps.user.models import User, UserSerializer
from common.response import ResponseSuccess, ResponseError, ResponseSuccessPage
from apps.role.models import RoleSerializer
from django.core.cache import cache
from utils.redisTool import set_cache
from openpyxl import Workbook
from Crypto.Cipher import AES
from common.const import SECRETKEY


# import logging
# 
# logger = logging.getLogger('django')


class UserViewSet(ViewSet):

    def list(self, request):
        params = json.loads(request.body)
        pageSize = params.get('pageSize')
        pageNo = params.get('pageNo')
        if not all([pageSize, pageNo]):
            return ResponseError(COMMON_RERROR.PAGENATE_PARAMS_IS_EMPTY)
        try:
            users = User.objects.filter(is_deleted=0)
            total = users.count()
            userLists = UserSerializer(Paginator(users, pageSize).page(pageNo), many=True).data
            return ResponseSuccessPage(data=userLists, total=total, pageSize=pageSize, pageNo=pageNo)
        except Exception as e:
            # logger.error(e)
            return ResponseError()

    def retrieve(self, request):
        params = json.loads(request.body)
        pageSize = params.get('pageSize')
        pageNo = params.get('pageNo')
        if not all([pageSize, pageNo]):
            return ResponseError(COMMON_RERROR.PAGENATE_PARAMS_IS_EMPTY)
        try:
            users = User.objects.filter(is_deleted=0)
            total = users.count()
            userLists = UserSerializer(Paginator(users, pageSize).page(pageNo), many=True).data
            return ResponseSuccessPage(data=userLists, total=total, pageSize=pageSize, pageNo=pageNo)
        except Exception as e:
            # logger.error(e)
            return ResponseError()

    def login(self, request):
        data = json.loads(request.body)
        email = data.get("email")
        aes_password = data.get('password')
        if email is None:
            return ResponseError(USER_RERROR.EMAIL_IS_REQUIRED)
        if aes_password is None:
            return ResponseError(USER_RERROR.PASSWORD_IS_REQUIRED)
        try:
            aesStr = AES.new(SECRETKEY.encode('utf-8'), AES.MODE_ECB)
            password = aesStr.decrypt(base64.b64decode(aes_password.encode('utf-8')))
            user = User.objects.get(email=email, password=password.decode('utf-8'))
        except User.DoesNotExist as e:
            return ResponseError(USER_RERROR.EMAIL_OR_PASSWORD_ERROR)
        try:
            jwt_payload_handler = api_settings.JWT_PAYLOAD_HANDLER
            jwt_encode_handler = api_settings.JWT_ENCODE_HANDLER
            payload = jwt_payload_handler(user)
            token = jwt_encode_handler(payload)
        except Exception as e:
            return ResponseError(USER_RERROR.EMAIL_OR_PASSWORD_ERROR)
        try:
            userInfos = UserSerializer(user).data
            roles = user.roles.values()
            menus = []
            buttons = []
            intefaces = []
            permissions = user.roles.prefetch_related('menus').all()
            userPermissions = RoleSerializer(permissions, many=True).data
            for userPermission in userPermissions:
                menu = userPermission.get('menus')
                button = userPermission.get('buttons')
                inteface = userPermission.get('intefaces')
                if (menu and menu not in menus):
                    menus = menus + menu
                if (button and button not in buttons):
                    buttons = buttons + button
                if (inteface and inteface not in intefaces):
                    intefaces = intefaces + inteface
            data = {**userInfos, 'token': token, 'roles': list(roles), 'menus': menus,
                    'buttons': buttons,
                    'intefaces': intefaces}
            set_cache(token, data)
            return ResponseSuccess(data=data)
        except Exception as e:
            print('sssss', e)
            return ResponseError()

    def register(self, request):
        pass

    def logout(self, request):
        try:
            token = request.META.get('HTTP_AUTHORIZATION')
            cache.delete(token)
            return ResponseSuccess()
        except Exception as e:
            return ResponseError()

    def create(self, request):
        data = json.loads(request.body)
        try:
            with transaction.atomic():
                user = User.objects.create(username=data['username'], password='1234@Abcd', avatar=data['avatar'],
                                           email=data['email'], phone_number=data['phone_number'],
                                           status=data['status'])
                user.roles.set(data.get('roles'))
            return ResponseSuccess()
        except Exception as e:
            # logger.error(e)
            return ResponseError()

    def update(self, request):
        data = json.loads(request.body)
        user_id = data.get('id')
        try:
            user = User.objects.get(id=user_id)
            with transaction.atomic():
                User.objects.filter(id=user_id).update(username=data['username'], avatar=data['avatar'],
                                                       phone_number=data['phone_number'],
                                                       email=data['email'], status=data['status'])
                user.roles.set(data.get('roles'))
            return ResponseSuccess()
        except Exception as e:
            # logger.error(e)
            return ResponseError()

    def destroy(self, request, user_id):
        if user_id is None:
            return ResponseError(USER_RERROR.USER_ID_IS_NOT_EXIST)
        try:
            user = User.objects.filter(id=user_id)
            user.delete()
            return ResponseSuccess()
        except Exception as e:
            # logger.error(e)
            return ResponseError()

    def details(self, request, user_id):
        if user_id is None:
            return ResponseError(USER_RERROR.USER_ID_IS_NOT_EXIST)
        try:
            user = User.objects.get(id=user_id)
            roles = user.roles.values_list('id', flat=True)
            userInfo = UserSerializer(user).data
            data = {**userInfo, "roles": list(roles)}
            return ResponseSuccess(data=data)
        except Exception as e:
            # logger.error(e)
            return ResponseError()

    def update_password(self, request):
        params = json.loads(request.body)
        if params['id'] is None:
            return JsonResponse(USER_RERROR.USER_IS_EMPTY)
        if params['password'] is None:
            return JsonResponse(USER_RERROR.PASSWORD_IS_EMPTY)
        try:
            user = User.objects.get(id=params['id'])
            user.password = params['password']
            user.save()
            return ResponseSuccess()
        except Exception as e:
            # logger.error(e)
            return ResponseError(USER_RERROR.USER_PASSWORD_UPDATE_FAILED)

    def update_avatar(self, request):
        params = json.loads(request.body)
        user_id = int(params['id'])
        if user_id is None:
            return JsonResponse(USER_RERROR.USER_IS_EMPTY)
        try:
            User.objects.filter(id=user_id).update(avatar=params['avatar'])
            return ResponseSuccess()
        except Exception as e:
            return ResponseError(USER_RERROR.USER_PASSWORD_UPDATE_FAILED)

    def export(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;filename=user.xlsx'
        # 创建Workbook和Sheet
        workbook = Workbook()
        worksheet = workbook.active
        # 写入表头
        worksheet.cell(row=1, column=1, value='username')
        worksheet.cell(row=1, column=2, value='email')
        worksheet.cell(row=1, column=3, value='phone_number')
        worksheet.cell(row=1, column=4, value='remark')
        # 写入数据
        users = User.objects.all()
        for i, user in enumerate(users, start=2):
            worksheet.cell(row=i, column=1, value=user.username)
            worksheet.cell(row=i, column=2, value=user.email)
            worksheet.cell(row=i, column=3, value=user.phone_number)
            worksheet.cell(row=i, column=4, value=user.remark)
        workbook.save(response)
        return response
