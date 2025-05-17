import json
import math
from datetime import datetime

from django.core.paginator import Paginator
from django.http import JsonResponse
from django.shortcuts import render, HttpResponse
from django.views import View
from rest_framework_jwt.settings import api_settings

from menu.models import SysRoleMenu, SysMenu
from role.models import SysRole, SysUserRole, SysRoleSerializer
from service_error.common import COMMON_RERROR
from service_error.user import USER_RERROR
from user.models import User, SysUserSerializer
from common.response import ResponseSuccess, ResponseError


class RegisterView(View):
    def post(self, request):
        pass


class LoginView(View):
    def post(self, request):
        params = json.loads(request.body)
        if params['email'] is None:
            return JsonResponse(USER_RERROR.USERNAME_IS_EMPTY)
        if params['password'] is None:
            return JsonResponse(USER_RERROR.PASSWORD_IS_EMPTY)
        email = params.get("email")
        password = params.get('password')
        if email is None or password is None:
            return ResponseError(USER_RERROR.USER_AND_PASSWORD_IS_REQUIRED)
        try:
            user = User.objects.get(email=email, password=password)
            jwt_payload_handler = api_settings.JWT_PAYLOAD_HANDLER
            jwt_encode_handler = api_settings.JWT_ENCODE_HANDLER
            payload = jwt_payload_handler(user)
            token = jwt_encode_handler(payload)
        except Exception as e:
            print(e)
            return ResponseError(USER_RERROR.USER_OR_PASSWORD_ERROR)
        roles = list(SysUserRole.objects.filter(user_id=user.id).all().values_list('role_id', flat=True))
        print("role_list", roles)
        menus = []
        rolesDetails = []
        for role in roles:
            menu = list(SysRoleMenu.objects.filter(role_id=role).all().values_list('menu_id', flat=True))
            menus = list(set(menus + menu))
            roleDetail = SysRole.objects.filter(id=role).all().values()
            rolesDetails = list(rolesDetails + list(roleDetail))
        menusDetails = []
        for menu in menus:
            detail = SysMenu.objects.filter(id=menu).all().values()
            menusDetails = list(menusDetails + list(detail))
        data = {**SysUserSerializer(user).data, 'token': token, 'roles': rolesDetails, 'menus': menusDetails}
        return ResponseSuccess(data=data)


class CreateView(View):
    # def get(self, request):
    #     return HttpResponse("获取用户列表")
    def post(self, request):
        user = json.loads(request.body)
        usr = User.objects.create(username=user['username'], password=user['password'], avatar=user['avatar'],
                                  email=user['email'], phone_number=user['phone_number'], status=user['status'])
        role_list = user['role']
        usrID = usr.id
        for roleId in role_list:
            print('roleId', roleId)
            SysUserRole.objects.create(user_id=usrID, role_id=roleId)
        return ResponseSuccess()


class SearchPageView(View):
    def post(self, request):
        params = json.loads(request.body)
        pageSize = params.get('pageSize')
        pageNo = params.get('pageNo')
        if not all([pageSize, pageNo]):
            return ResponseError(COMMON_RERROR.PAGENATE_PARAMS_IS_EMPTY)
        userLists = Paginator(User.objects.filter(is_deleted=0), pageSize).page(pageNo)
        total = User.objects.filter(is_deleted=0).count()
        users = SysUserSerializer(userLists.object_list.values(), many=True).data
        # roles = list(SysUserRole.objects.filter(user=user_id).all().values_list('role_id', flat=True))
        data = {'lists': users, 'total': total, 'pageSize': pageSize, 'pageNo': pageNo}
        return ResponseSuccess(data=data)


class SearchListsView(View):
    def post(self, request):
        return ResponseSuccess()


class UpdatePasswordView(View):
    def post(self, request):
        params = json.loads(request.body)
        if params['id'] is None:
            return JsonResponse(USER_RERROR.USER_IS_EMPTY)
        if params['password'] is None:
            return JsonResponse(USER_RERROR.PASSWORD_IS_EMPTY)
        try:
            user = User.objects.get(id=params['id'])
            user.password = params['password']
            user.save()
        except Exception as e:
            return ResponseError(USER_RERROR.USER_PASSWORD_UPDATE_FAILED)
        return ResponseSuccess()


class UpdateView(View):
    def post(self, request):
        data = json.loads(request.body)
        id = data.get('id')
        user = User.objects.filter(id=id).update(username=data['username'], password='1234@Abcd',
                                                 avatar=data['avatar'], phone_number=data['phone_number'],
                                                 email=data['email'], status=data['status'])
        SysUserRole.objects.filter(user_id=id).delete()
        for roleId in data['role']:
            SysUserRole.objects.update_or_create(user_id=id, role_id=roleId)
        return ResponseSuccess()


class DetailView(View):
    def post(self, request, user_id):
        if user_id is None:
            return ResponseError(USER_RERROR.USER_ID_IS_NOT_EXIST)
        user = User.objects.get(id=user_id)
        users = SysUserSerializer(user).data
        roles = list(SysUserRole.objects.filter(user=user_id).all().values_list('role_id', flat=True))
        data = {**users, 'role': roles}
        return ResponseSuccess(data=data)


class DeleteView(View):
    def delete(self, request, user_id):
        if user_id is None:
            return ResponseError(USER_RERROR.USER_ID_IS_NOT_EXIST)
        user = User.objects.get(id=user_id)
        user.is_deleted = 1
        user.save()
        SysUserRole.objects.filter(user_id=user_id).delete()
        return ResponseSuccess()


class LogoutView(View):
    def post(self, request):
        return ResponseSuccess()


class GetAssetsView(View):
    def get(self, request):
        page_size = request.GET.get("pageSize")
        page_no = request.GET.get("pageNo")
        datas = []
        for i in range(1, int(page_size)):
            item_dict = {
                'id': math.radians(),
                'title': "椅子",
                'categopry': "办公用品",
                'department': "人事行政部",
                'number': "FZC00",
                'status': 0,
                'assetsStatus': 1,
                'belong': "sutter",
                'updator': "Tom",
                'update': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
            datas.append(item_dict)
        print(datas)
        return JsonResponse({'code': 200, 'data': json.dumps(datas), 'msg': 'success'})


from openpyxl import Workbook


class ExportView(View):
    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment;filename=user.xlsx'
        # 创建Workbook和Sheet
        workbook = Workbook()
        worksheet = workbook.active
        # 写入表头
        worksheet.cell(row=1, column=1, value='username')
        worksheet.cell(row=1, column=2, value='email')
        worksheet.cell(row=1, column=3, value='phone_number')
        worksheet.cell(row=1, column=4, value='remark')
        # 写入数据
        users = User.objects.all()
        for i, user in enumerate(users, start=2):
            worksheet.cell(row=i, column=1, value=user.username)
            worksheet.cell(row=i, column=2, value=user.email)
            worksheet.cell(row=i, column=3, value=user.phone_number)
            worksheet.cell(row=i, column=4, value=user.remark)
        workbook.save(response)
        return response


def ExportExcel():
    pass
